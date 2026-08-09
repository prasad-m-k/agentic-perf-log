#!/usr/bin/env python3
"""
Append one row to the Log tab of agent-correction-tracker.xlsx.

Meant to be called by an agent (Claude Code, or any CLI-driven coding
assistant) right after a task is reviewed, so the tracker fills itself in
instead of relying on a human to remember.

Usage (v1 fields, still required):
    python3 update_log.py \
        --task-id TASK-042 \
        --category "Bug Fix" \
        --agent "Claude Code" \
        --corrected yes \
        --correction-type "Missed Edge Case" \
        --notes "Off-by-one on empty list input, human fixed before merge" \
        --logged-by agent

Usage (v2 fields, all optional, blank if omitted):
    python3 update_log.py \
        --task-id TASK-042 --category "Bug Fix" --corrected yes \
        --correction-type "Missed Edge Case" \
        --risk-tier Medium \
        --spec-completeness 4 \
        --prompt-turns 2 \
        --review-time 8 \
        --est-manual-time 25 \
        --diff-retention 78.5 \
        --test-modified no \
        --non-goal-violation no

Valid --category values:
    Test Generation, New Feature, Refactor, Bug Fix, Documentation,
    Config/Boilerplate, Other

Valid --correction-type values (only used when --corrected yes):
    Wrong Assumption, Missed Edge Case, Scope Creep, Style Mismatch,
    Logic Error, Other

Valid --risk-tier values:
    Low, Medium, High

v2 field notes:
- --spec-completeness is a 1-5 self-rating of how complete the task spec
  was before the agent started (5 = Goal/Boundaries/Edge cases/Output
  format/Non-goals all present).
- --review-time and --est-manual-time are both in minutes. Their ratio
  is the Review-to-Authored Ratio on the Dashboard tab. Above 1.0 means
  reviewing the output took longer than writing it by hand would have.
- --diff-retention is a percentage (0-100): how much of the agent's
  generated diff survived into the final commit.
- --test-modified flags whether the agent edited or relaxed an existing
  test to make its own code pass. This is the one to watch closest.
- --non-goal-violation flags whether the agent edited files or scope
  outside what the task spec allowed.
- All v2 fields default to blank if omitted, so v1-style calls keep
  working unchanged.

Notes:
- This appends to the underlying spreadsheet directly. It does not run
  the formula recalculation, most spreadsheet apps (Excel, LibreOffice,
  Google Sheets after import) recalc automatically on open.
- If you're running this outside a repo with the xlsx already present,
  point --file at its path.
- Safe for multiple agents writing to the same file at once. Each call
  takes an exclusive lock on the workbook first, so concurrent writes
  queue up instead of clobbering each other. No extra install needed.
"""

import argparse
import contextlib
import datetime
import os
import sys
import time

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

VALID_CATEGORIES = {
    "Test Generation", "New Feature", "Refactor", "Bug Fix",
    "Documentation", "Config/Boilerplate", "Other",
}
VALID_TYPES = {
    "Wrong Assumption", "Missed Edge Case", "Scope Creep",
    "Style Mismatch", "Logic Error", "Other",
}
VALID_TIERS = {"Low", "Medium", "High"}


@contextlib.contextmanager
def file_lock(target_path, timeout=30, poll_interval=0.2):
    """
    Simple cross-platform exclusive lock so multiple agents writing to the
    same workbook at the same time queue up instead of one overwriting
    the other's row. Uses a sibling .lock file and atomic O_CREAT|O_EXCL,
    no third-party dependency required.
    """
    lock_path = target_path + ".lock"
    deadline = time.time() + timeout
    fd = None
    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            break
        except FileExistsError:
            if time.time() > deadline:
                raise TimeoutError(
                    f"Could not acquire lock on {target_path} after {timeout}s. "
                    f"Another process may be stuck holding {lock_path}. "
                    f"Delete it manually if you're sure nothing else is writing."
                )
            time.sleep(poll_interval)
    try:
        os.write(fd, str(os.getpid()).encode())
        yield
    finally:
        os.close(fd)
        try:
            os.remove(lock_path)
        except FileNotFoundError:
            pass


def main():
    p = argparse.ArgumentParser(description="Append a row to the agent correction Log tab.")
    p.add_argument("--file", default="agent-correction-tracker.xlsx", help="Path to the workbook")
    p.add_argument("--task-id", required=True)
    p.add_argument("--category", required=True, choices=sorted(VALID_CATEGORIES))
    p.add_argument("--agent", default="Claude Code")
    p.add_argument("--corrected", required=True, choices=["yes", "no"])
    p.add_argument("--correction-type", default="", choices=[""] + sorted(VALID_TYPES))
    p.add_argument("--notes", default="")
    p.add_argument("--logged-by", default="agent")
    p.add_argument("--date", default=None, help="YYYY-MM-DD, defaults to today")
    p.add_argument("--lock-timeout", type=float, default=30, help="Seconds to wait for the file lock before giving up")

    # v2 fields, all optional
    p.add_argument("--risk-tier", default="", choices=[""] + sorted(VALID_TIERS))
    p.add_argument("--spec-completeness", type=int, default=None, choices=[None, 1, 2, 3, 4, 5])
    p.add_argument("--prompt-turns", type=int, default=None)
    p.add_argument("--review-time", type=float, default=None, help="Minutes spent reviewing the output")
    p.add_argument("--est-manual-time", type=float, default=None, help="Estimated minutes to write it by hand")
    p.add_argument("--diff-retention", type=float, default=None, help="Percent of agent diff kept in final commit, 0-100")
    p.add_argument("--test-modified", default="", choices=["", "yes", "no"])
    p.add_argument("--non-goal-violation", default="", choices=["", "yes", "no"])
    args = p.parse_args()

    if args.corrected == "yes" and not args.correction_type:
        sys.exit("error: --correction-type is required when --corrected yes")
    if args.corrected == "no" and args.correction_type:
        sys.exit("error: --correction-type should be empty when --corrected no")

    date = args.date or datetime.date.today().isoformat()

    try:
        with file_lock(args.file, timeout=args.lock_timeout):
            wb = openpyxl.load_workbook(args.file)
            log = wb["Log"]

            # Append after the last row that actually has data in the Task ID
            # column. Bounded to row 500 so we never wander into the legend text
            # block that sits below the pre-formatted rows on the shipped sheet.
            row = 1
            for r in range(2, 500):
                if log.cell(row=r, column=2).value not in (None, ""):
                    row = r
            row += 1

            def yn(val):
                if val == "yes":
                    return "Y"
                if val == "no":
                    return "N"
                return ""

            values = [
                date,
                args.task_id,
                args.category,
                args.agent,
                "Y" if args.corrected == "yes" else "N",
                args.correction_type,
                args.notes,
                args.logged_by,
                args.risk_tier,
                args.spec_completeness if args.spec_completeness is not None else "",
                args.prompt_turns if args.prompt_turns is not None else "",
                args.review_time if args.review_time is not None else "",
                args.est_manual_time if args.est_manual_time is not None else "",
                args.diff_retention if args.diff_retention is not None else "",
                yn(args.test_modified),
                yn(args.non_goal_violation),
            ]
            for col, v in enumerate(values, start=1):
                log.cell(row=row, column=col, value=v)

            wb.save(args.file)
            print(f"logged row {row}: {args.task_id} ({args.category}, corrected={args.corrected})")
    except TimeoutError as e:
        sys.exit(f"error: {e}")


if __name__ == "__main__":
    main()
